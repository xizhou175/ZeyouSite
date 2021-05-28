from django.shortcuts import render
from random import randint
from django.core.mail import send_mail
from rest_framework.generics import CreateAPIView
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import JsonResponse, QueryDict
from django.forms.models import model_to_dict
from .models import User
from students.models import Student
from academies.models import Academy
from .serializers import CreateUserSerializer
import json
from datetime import date, datetime
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from django.db.models import Q
import os
import openpyxl
import xlrd


class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, date):
            return obj.strftime("%Y-%m-%d")
        else:
            return json.JSONEncoder.default(self, obj)


# Create your views here.
class UserView(CreateAPIView):
    """
    用户注册
    """
    serializer_class = CreateUserSerializer


class UserUpdateView(APIView):
    """
    用户详情
    """

    def post(self, request):

        user = User.objects.filter(id=request.data["user_id"]).first()
        if user is None:
            return Response({'message': "User not found"})

        # noinspection PyBroadException
        try:
            User.objects.filter(id=request.data["user_id"]).update(phone=request.data['phone'],
                                                                   email=request.data['email'],
                                                                   username=request.data["username"])
            # 修改成功
            # Successfully modified
            return Response({'message': "Successfully modified", 'status': 200})
        except Exception as e:
            # 参数不全
            # Incomplete parameters
            return Response({'message': "Incomplete parameters", 'status': 400})


class UpdatePwdView(APIView):

    def post(self, request):

        data = request.data
        user = User.objects.get(id=data['user_id'])

        if user is None:
            return Response({'message': "User not found"})
        if not user.check_password(data['old_password']):
            # The original password was entered incorrectly 原密码输入有误
            return Response({'message': "The original password was entered incorrectly "})
        if data['new_password'] != data['confirm_password']:
            # The passwords do not match twice, please re-enter 两次密码不一致，请重新输入
            return Response('The passwords do not match twice, please re-enter')

        user.set_password(data['new_password'])
        user.save()
        # Successfully modified 修改成功
        return Response({'message': "Successfully modified", "status": 200})


class FilterDepartmentView(APIView):

    def post(self, request):
        department_list = ['little_assistant', 'consultant', 'service_consultant', 'paper_writer']
        data = request.data
        department = data['department']
        name = data['username']
        pageNum = int(data['pageNum'])
        click = int(data['click'])

        # search_dict = {"little_assistant": name, "consultant": name, "service_consultant": name, "paper_writer": name}
        if click == 1 or click == 2:
            info = Student.objects.filter(Q(little_assistant=name) | Q(consultant=name) | Q(service_consultant=name)
                                          | Q(paper_writer=name)).all()

        if click == 3:
            info = Academy.objects.filter(name=name).all()
        ret = []
        for i in range(info.count()):
            data1 = model_to_dict(info[i])
            data1["status"] = 200
            if click == 1:
                data2 = dict()
                data2["name"] = data1["name"]
                data2["graduation_date"] = data1["graduation_date"]
                data2["school"] = data1["school"]
                data2["status"] = 200
                ret.append(data2)
                continue
            ret.append(data1)
        jsonArr = json.dumps(ret, ensure_ascii=False, cls=DateEncoder)
        return JsonResponse(jsonArr, safe=False)


class UploadView(APIView):
    def post(self, request):
        file = request.FILES['myfile']
        localFile = open('D:\\PythonProject\\Zeyou\\file\\' + file.name, 'wb+')
        for chunk in file.chunks():
            localFile.write(chunk)
        localFile.close()
        return Response({'message': "Successfully uploaded", "status": 48})


class ParseExcelView(APIView):
    def post(self, request):
        if os.path.exists('D:\\PythonProject\\Zeyou\\file\\dataset.xlsx') == False:
            return Response({'message': "No such file to be parsed", "status": 400})

        wb = openpyxl.load_workbook('D:\\PythonProject\\Zeyou\\file\\dataset.xlsx', data_only=True)
        allsheets = wb.sheetnames
        #print(len(allsheets))
        index = 0
        while index < 2:
            print(index)
            sheet = wb[allsheets[index]]
            maxrow = sheet.max_row
            maxcol = sheet.max_column
            row_dict = {}

            if index == 0:
                #continue
                headers = ["customer_state", "source", "date_to_add", "name", "gender", "wechat_num", "area",
                           "phone", "little_assistant", "consultant", "service_consultant", "paper_writer", "identity",
                           "school", "school_type", "curriculum_system", "curriculum_system_note", "graduation_date",
                           "application_level", "major", "target_country", "GPA", "TOEFL", "IELTS", "SAT", "ACT", "GRE"]
                lists = []
                for row in range(2, maxrow + 1):
                    r = {}
                    for col in range(1, len(headers) + 1):
                        key = headers[col - 1]
                        r[key] = sheet.cell(row=row, column=col + 1).value
                        if key == 'gender':
                            if r[key] == '男':
                                r[key] = 'M'
                            else:
                                r[key] = 'F'

                        if key == 'identity':
                            if r[key] == '家长':
                                r[key] = 0
                            else:
                                r[key] = 1

                        if key == 'customer_state':
                            if r[key] == '未分配未购买':
                                r[key] = 0
                            if r[key] == '已分配未购买':
                                r[key] = 1
                            if r[key] == '未分配已购买':
                                r[key] = 2
                            if r[key] == '已分配已购买':
                                r[key] = 3
                            if r[key] == '已签约未购买':
                                r[key] = 4
                            if r[key] == '已签约已购买':
                                r[key] = 5

                        if key == 'date_to_add':
                            if r[key] is not None:
                                dtime = xlrd.xldate_as_tuple(r[key], 0)  # 转化为元组形式
                                d = date(dtime[0], dtime[1], dtime[2])
                                r[key] = d

                        if r[key] is None:
                            r[key] = ""
                    lists.append(r)

                sqllist = []
                for cell in lists:
                    # for header in headers:
                    customer_state = cell['customer_state']
                    source = cell['source']
                    date_to_add = cell['date_to_add']
                    name = cell['name']
                    gender = cell['gender']
                    wechat_num = cell['wechat_num']
                    area = cell['area']
                    phone = cell['phone']
                    little_assistant=cell["little_assistant"]
                    consultant = cell["consultant"]
                    service_consultant = cell["service_consultant"]
                    paper_writer = cell["paper_writer"]
                    identity = cell['identity']
                    school=cell["school"]
                    school_type=cell["school_type"]
                    curriculum_system = cell["curriculum_system"]
                    curriculum_system_note = cell["curriculum_system_note"]
                    application_level = cell["application_level"]
                    major = cell["major"]
                    target_country = cell["target_country"]
                    GPA = cell["GPA"]
                    TOEFL=cell["TOEFL"]
                    IELTS=cell["IELTS"]
                    SAT = cell["SAT"]
                    ACT = cell["ACT"]
                    GRE = cell["GRE"]
                    sql = Student(customer_state=customer_state, date_to_add=date_to_add, source=source, name=name,
                                  wechat_num=wechat_num, area=area, phone=phone, gender=gender, identity=identity,
                                  little_assistant=little_assistant, consultant=consultant, service_consultant=service_consultant,
                                  paper_writer=paper_writer, school=school, school_type=school_type,
                                  curriculum_system=curriculum_system, curriculum_system_note=curriculum_system_note,
                                  application_level=application_level, major=major, target_country=target_country,
                                  GPA=GPA, TOEFL=TOEFL, IELTS=IELTS, SAT=SAT, ACT=ACT, GRE=GRE)
                    sqllist.append(sql)
                try:
                    Student.objects.bulk_create(sqllist)
                except:
                    return Response({"message": "Parameters are incomplete and cannot be saved", "status": 400})
                index += 1

            elif index == 1:

                headers = ["name", "source", "product_type", "teaching_assistant", "sales", "teacher",
                           "date_of_purchasing", "product", "date_of_lecture", "hours_of_lecture", "price_per_hour"
                           , "price_overall", "cur_state"]
                lists = []
                for row in range(2, maxrow + 1):
                    r = {}
                    for col in range(1, len(headers) + 1):
                        key = headers[col - 1]
                        r[key] = sheet.cell(row=row, column=col).value

                        if key == 'hours_of_lecture' or key == 'price_per_hour' or key == 'price_overall':
                            print(r[key])
                            if r[key] is not None:
                                r[key] = int(r[key])

                        if key == 'date_of_lecture' or key == 'date_of_purchasing':
                            if r[key] is not None:
                                dtime = xlrd.xldate_as_tuple(r[key], 0)  # 转化为元组形式
                                d = date(dtime[0], dtime[1], dtime[2])
                                r[key] = d

                            #else:
                            #    r[key] = date(1000, 1, 1)

                        if r[key] is None and key != 'date_of_lecture' and key != 'date_of_purchasing':
                            r[key] = ""
                    lists.append(r)

                sqllist = []
                for cell in lists:
                    # for header in headers:
                    name = cell['name']
                    source = cell['source']
                    date_of_purchasing = cell['date_of_purchasing']
                    hours_of_lecture = cell['hours_of_lecture']
                    price_overall = cell["price_overall"]
                    date_of_lecture = cell['date_of_lecture']
                    product_type = cell['product_type']
                    product = cell['product']
                    price_per_hour = cell['price_per_hour']
                    cur_state = cell['cur_state']
                    teacher = cell['teacher']
                    teaching_assistant = cell['teaching_assistant']
                    sales = cell['sales']
                    sql = Academy(name=name, source=source, price_overall=price_overall, date_of_purchasing=date_of_purchasing,
                                  date_of_lecture=date_of_lecture, hours_of_lecture=hours_of_lecture, product=product,
                                  product_type=product_type, price_per_hour=price_per_hour, cur_state=cur_state, teacher=teacher,
                                  teaching_assistant=teaching_assistant, sales=sales)
                    sqllist.append(sql)
                try:
                    Academy.objects.bulk_create(sqllist)
                except:
                    return Response({"message": "Parameters are incomplete and cannot be saved", "status": 400})
                index += 1
        return Response({'message': "Successfully parsed", "status": 200})
